# -*- coding: utf-8 -*-
"""
نظام تحليل نتائج الاختبارات - ثيم وزارة التعليم
-------------------------------------------------
تطبيق Flask يقرأ ملف درجات (Excel/CSV)، يحسب النسب المئوية والتقديرات،
ويرتب الطلاب على مستوى الفصل وعلى مستوى الصف، مع إحصائيات وتصدير للنتائج.

صيغ الملف المدعومة:
  1. شيت واحد يحتوي أعمدة: اسم الطالب، الفصل، الدرجة
  2. شيت لكل فصل (اسم الشيت = اسم الفصل) يحتوي أعمدة: اسم الطالب، الدرجة
يمكن الجمع بين الصيغتين في الملف نفسه؛ إذا وُجد عمود الفصل داخل الشيت فهو الأولوية.

الترتيب: متسلسل (1، 2، 3...) حسب النسبة تنازليًا، وعند التساوي حسب الاسم أبجديًا.

التشغيل محليًا:
  python app.py
"""
import io
import os
import re

import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # الحد الأقصى لحجم الملف المرفوع: 10MB

# التقديرات حسب نموذج تحليل النتائج المعتمد في وزارة التعليم
LEVEL_NAMES = ["ممتاز", "جيد جدًا", "جيد", "مقبول", "ضعيف"]
DEFAULT_THRESHOLDS = (("t_excellent", 90), ("t_vgood", 80), ("t_good", 70), ("t_pass", 50))

# الأسماء المقبولة لكل عمود (عربي/إنجليزي)، تُطابق بعد التوحيد
COLUMN_ALIASES = {
    "name": ["اسم الطالب", "اسم الطالبة", "الاسم", "اسم", "الطالب", "الطالبة", "name", "student", "student name"],
    "class": ["الفصل", "فصل", "الشعبة", "شعبة", "class", "section"],
    "score": ["الدرجة", "درجة", "الدرجه", "درجه", "المجموع", "score", "grade", "mark", "total"],
}

ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
SHEET_TITLE_INVALID = re.compile(r"[\\/*?:\[\]]")


# ---------------------------------------------------------------------------
# أدوات مساعدة
# ---------------------------------------------------------------------------
def normalize(text):
    """توحيد النص العربي لأغراض المقارنة (إزالة التشكيل، توحيد الهمزات والتاء المربوطة)."""
    text = str(text).strip().lower()
    text = re.sub(r"[\u064B-\u0652\u0640]", "", text)
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    text = text.replace("ة", "ه").replace("ى", "ي")
    return text


def find_column(df, key):
    """إرجاع اسم العمود المطابق للمفتاح المطلوب، أو None إذا لم يوجد."""
    aliases = [normalize(a) for a in COLUMN_ALIASES[key]]
    for col in df.columns:
        if normalize(col) in aliases:
            return col
    for col in df.columns:
        n = normalize(col)
        if any(a in n for a in aliases):
            return col
    return None


def to_number(value):
    """تحويل القيمة إلى رقم مع دعم الأرقام العربية والفواصل، أو None إذا تعذر."""
    if value is None or pd.isna(value):
        return None
    s = str(value).translate(ARABIC_DIGITS).replace("٫", ".").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def level_of(pct, thresholds):
    """تحديد التقدير بناءً على النسبة المئوية وحدود التقديرات (تنازلية)."""
    for threshold, name in zip(thresholds, LEVEL_NAMES):
        if pct >= threshold:
            return name
    return LEVEL_NAMES[-1]


def level_ranges(thresholds, total):
    """نطاق الدرجات الخام لكل تقدير، بالشكل المستخدم في نموذج الوزارة (مثال: 17.1 - 19)."""
    bounds = [round(t / 100 * total, 2) for t in thresholds]
    ranges = {}
    upper = total
    for name, low in zip(LEVEL_NAMES[:-1], bounds):
        ranges[name] = {"low": low, "high": upper}
        upper = round(low - 0.01, 2)
    ranges[LEVEL_NAMES[-1]] = {"low": 0, "high": upper}
    return ranges


def group_stats(df, pass_threshold, total):
    """إحصائيات وصفية لمجموعة من الطلاب (فصل أو الصف كامل)."""
    pct = df["pct"]
    score = df["score"]
    n = len(df)
    return {
        "count": int(n),
        "avg": round(float(pct.mean()), 2) if n else 0,
        "median": round(float(pct.median()), 2) if n else 0,
        "std": round(float(pct.std(ddof=0)), 2) if n > 1 else 0,
        "max": round(float(pct.max()), 2) if n else 0,
        "min": round(float(pct.min()), 2) if n else 0,
        "score_max": round(float(score.max()), 2) if n else 0,
        "score_min": round(float(score.min()), 2) if n else 0,
        "score_avg": round(float(score.mean()), 2) if n else 0,
        "score_sum": round(float(score.sum()), 2) if n else 0,
        # نسبة التحصيل = مجموع الدرجات ÷ (عدد الطلاب × الدرجة الكلية)
        "attainment": round(float(score.sum() / (n * total) * 100), 2) if n else 0,
        "pass_rate": round(float((pct >= pass_threshold).mean() * 100), 1) if n else 0,
        "levels": {name: int((df["level"] == name).sum()) for name in LEVEL_NAMES},
    }


def read_sheets(file):
    """قراءة الملف المرفوع وإرجاع قائمة (اسم الشيت، DataFrame). ملفات CSV تُعامل كشيت واحد."""
    if file.filename.lower().endswith(".csv"):
        return [("الدرجات", pd.read_csv(file))]
    sheets = pd.read_excel(file, sheet_name=None)
    return [(name, df) for name, df in sheets.items() if not df.dropna(how="all").empty]


def build_dataset(sheets):
    """
    دمج الشيتات في جدول واحد بالأعمدة: name, class, score.
    - إذا احتوى الشيت على عمود الفصل، يُستخدم.
    - وإلا يُعتبر اسم الشيت هو اسم الفصل.
    يرجع (DataFrame, قائمة أخطاء).
    """
    frames, errors = [], []
    for sheet_name, df in sheets:
        df = df.dropna(how="all")
        df.columns = [str(c).strip() for c in df.columns]
        name_col = find_column(df, "name")
        score_col = find_column(df, "score")
        class_col = find_column(df, "class")

        missing = [label for label, col in (("اسم الطالب", name_col), ("الدرجة", score_col)) if col is None]
        if missing:
            errors.append(f"الشيت «{sheet_name}»: الأعمدة الناقصة ({'، '.join(missing)}). "
                          f"الأعمدة الموجودة: {'، '.join(df.columns)}")
            continue

        class_values = (df[class_col].astype(str).str.strip() if class_col is not None
                        else pd.Series([str(sheet_name).strip()] * len(df), index=df.index))

        frames.append(pd.DataFrame({
            "name": df[name_col].astype(str).str.strip(),
            "class": class_values.str.translate(ARABIC_DIGITS),
            "score": df[score_col].map(to_number),
        }))

    if not frames:
        return pd.DataFrame(columns=["name", "class", "score"]), errors
    return pd.concat(frames, ignore_index=True), errors


def style_sheet(ws, headers, rows):
    """كتابة جدول منسق في ورقة عمل: رؤوس ملونة، حدود، محاذاة، اتجاه من اليمين لليسار."""
    head_fill = PatternFill("solid", fgColor="1E9E8A")
    head_font = Font(bold=True, color="FFFFFF", name="Arial")
    thin = Side(style="thin", color="C8D3CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment, cell.border = head_fill, head_font, center, border
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment, cell.border = center, border
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(header)) + 6)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# المسارات
# ---------------------------------------------------------------------------
LOGO_DIRS = ("static", ".")
LOGO_EXTENSIONS = (".png", ".jpg", ".jpeg", ".svg", ".webp", ".gif")


def find_logo():
    """البحث عن ملف شعار باسم logo بأي امتداد أو حالة أحرف، في مجلد static أو مجلد المشروع."""
    base = os.path.dirname(os.path.abspath(__file__))
    for folder in LOGO_DIRS:
        path = os.path.join(base, folder)
        if not os.path.isdir(path):
            continue
        for entry in sorted(os.listdir(path)):
            stem, ext = os.path.splitext(entry)
            if stem.lower() == "logo" and ext.lower() in LOGO_EXTENSIONS:
                return os.path.join(path, entry)
    return None


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/logo")
def logo():
    """إرجاع ملف الشعار إن وُجد؛ وإلا 404 ليعرض القالب النص البديل."""
    path = find_logo()
    if not path:
        return "", 404
    return send_file(path)


@app.route("/analyze", methods=["POST"])
def analyze():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify(error="لم يتم رفع ملف الدرجات."), 400

    total = to_number(request.form.get("total", ""))
    if not total or total <= 0:
        return jsonify(error="الدرجة الكلية يجب أن تكون رقمًا أكبر من صفر."), 400

    thresholds = [to_number(request.form.get(k, d)) for k, d in DEFAULT_THRESHOLDS]
    if any(t is None for t in thresholds) or thresholds != sorted(thresholds, reverse=True):
        return jsonify(error="حدود التقديرات غير صحيحة؛ يجب أن تكون تنازلية "
                             "(ممتاز > جيد جدًا > جيد > مقبول)."), 400

    support_threshold = to_number(request.form.get("support", 60)) or 60

    try:
        sheets = read_sheets(file)
    except Exception:
        return jsonify(error="تعذّرت قراءة الملف. الصيغ المدعومة: Excel (.xlsx) أو CSV."), 400
    if not sheets:
        return jsonify(error="الملف لا يحتوي على أي بيانات."), 400

    data, errors = build_dataset(sheets)
    if data.empty:
        return jsonify(error=" | ".join(errors) or "لا توجد بيانات صالحة في الملف."), 400

    # استبعاد الصفوف بدون اسم أو بدرجة غير رقمية
    invalid_mask = data["score"].isna() | data["name"].isin(["", "nan", "None"])
    skipped = int(invalid_mask.sum())
    data = data[~invalid_mask].reset_index(drop=True)
    if data.empty:
        return jsonify(error="لا توجد صفوف صالحة؛ يجب أن يحتوي عمود الدرجة على أرقام."), 400

    over = data[data["score"] > total]
    if not over.empty:
        return jsonify(error=f"يوجد {len(over)} طالب درجته أعلى من الدرجة الكلية ({total:g})، "
                             f"مثال: {over.iloc[0]['name']} = {over.iloc[0]['score']:g}."), 400

    # الحسابات الأساسية
    data["pct"] = (data["score"] / total * 100).round(2)
    data["level"] = data["pct"].map(lambda p: level_of(p, thresholds))

    # ترتيب متسلسل بدون تكرار: النسبة تنازليًا ثم الاسم أبجديًا
    data = data.sort_values(["pct", "name"], ascending=[False, True]).reset_index(drop=True)
    data["rank_grade"] = range(1, len(data) + 1)
    data["rank_class"] = data.groupby("class").cumcount() + 1

    classes = sorted(data["class"].unique().tolist())
    grade_stats = group_stats(data, thresholds[3], total)
    class_stats = {}
    for c in classes:
        st = group_stats(data[data["class"] == c], thresholds[3], total)
        st["gap"] = round(st["avg"] - grade_stats["avg"], 2)  # الفرق عن متوسط الصف
        class_stats[c] = st

    # توزيع النسب على فئات بعرض 10% (الفئة الأخيرة تشمل 100)
    bins = list(range(0, 100, 10))
    labels = [f"{b}–{b + 9}" for b in bins]
    labels[-1] = "90–100"

    def in_bin(series, b):
        return (series >= b) & ((series < b + 10) | (b == 90))

    histogram = {
        "labels": labels,
        "grade": [int(in_bin(data["pct"], b).sum()) for b in bins],
        "classes": {c: [int(in_bin(data.loc[data["class"] == c, "pct"], b).sum()) for b in bins]
                    for c in classes},
    }

    needs_support = data[data["pct"] < support_threshold].sort_values("pct")
    students = data[["name", "class", "score", "pct", "level", "rank_class", "rank_grade"]].to_dict("records")

    return jsonify({
        "total": total,
        "thresholds": thresholds,
        "ranges": level_ranges(thresholds, total),
        "support_threshold": support_threshold,
        "levels": LEVEL_NAMES,
        "classes": classes,
        "students": students,
        "grade_stats": grade_stats,
        "class_stats": class_stats,
        "histogram": histogram,
        "top10": students[:10],
        "needs_support": needs_support[["name", "class", "score", "pct", "level"]].to_dict("records"),
        "skipped": skipped,
        "warnings": errors,
    })


@app.route("/export", methods=["POST"])
def export():
    """تصدير النتائج إلى ملف Excel: ترتيب الصف، شيت لكل فصل، الإحصائيات، الأوائل، والطلاب الذين يحتاجون متابعة."""
    payload = request.get_json(silent=True) or {}
    students = payload.get("students", [])
    classes = payload.get("classes", [])
    class_stats = payload.get("class_stats", {})
    grade_stats = payload.get("grade_stats", {})
    total = payload.get("total", "")
    if not students:
        return jsonify(error="لا توجد نتائج للتصدير."), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "ترتيب الصف"
    style_sheet(ws,
                ["م", "اسم الطالب", "الفصل", "الدرجة", f"النسبة % (من {total})", "التقدير"],
                [[s["rank_grade"], s["name"], s["class"], s["score"], s["pct"], s["level"]] for s in students])

    for c in classes:
        ws = wb.create_sheet(SHEET_TITLE_INVALID.sub("-", f"فصل {c}")[:31])
        rows = sorted((s for s in students if s["class"] == c), key=lambda s: s["rank_class"])
        style_sheet(ws,
                    ["م", "اسم الطالب", "الدرجة", "النسبة %", "التقدير", "الترتيب على الصف"],
                    [[s["rank_class"], s["name"], s["score"], s["pct"], s["level"], s["rank_grade"]] for s in rows])

    ws = wb.create_sheet("الإحصائيات")
    headers = ["المجموعة", "عدد الطلاب", "أعلى درجة", "أقل درجة", "متوسط الدرجات", "نسبة التحصيل %",
               "مجموع الدرجات", "نسبة الاجتياز %"] + LEVEL_NAMES
    rows = []
    groups = [(f"فصل {c}", class_stats.get(c, {})) for c in classes] + [("الصف كامل", grade_stats)]
    for label, st in groups:
        rows.append([label, st.get("count"), st.get("score_max"), st.get("score_min"), st.get("score_avg"),
                     st.get("attainment"), st.get("score_sum"), st.get("pass_rate")]
                    + [st.get("levels", {}).get(l, 0) for l in LEVEL_NAMES])
    style_sheet(ws, headers, rows)

    top10 = payload.get("top10", [])
    style_sheet(wb.create_sheet("الأوائل"),
                ["م", "اسم الطالب", "الفصل", "الدرجة", "النسبة %"],
                [[s["rank_grade"], s["name"], s["class"], s["score"], s["pct"]] for s in top10])

    support = payload.get("needs_support", [])
    threshold = payload.get("support_threshold", "")
    style_sheet(wb.create_sheet("يحتاجون متابعة"),
                ["اسم الطالب", "الفصل", "الدرجة", f"النسبة % (أقل من {threshold})", "التقدير"],
                [[s["name"], s["class"], s["score"], s["pct"], s["level"]] for s in support])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="نتائج_التحليل.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/template")
def template():
    """نموذج Excel جاهز: شيت لكل فصل، واسم الشيت هو اسم الفصل."""
    wb = Workbook()
    wb.remove(wb.active)
    for cls, rows in (("6-أ", [["نورة محمد", 45], ["سارة فهد", 42]]),
                      ("6-ب", [["ريم خالد", 38], ["هند عمر", 30]]),
                      ("6-ج", [["لمى ناصر", 47], ["جود علي", 29]])):
        style_sheet(wb.create_sheet(cls), ["اسم الطالب", "الدرجة"], rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="نموذج_الدرجات.xlsx")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
